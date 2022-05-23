import openpyxl

# Load workbook with path
wb = openpyxl.load_workbook("C:\\Users\\Asus\\workspace_python\\seleniumIntro\\Screenshots\\TestData.xlsx")
# Identify active worksheet/ worksheet choose
s = wb.active
# Get max row & column
max_row_count = s.max_row
max_col_count = s.max_column
print("Current sheet: " + s.title)
print("Total row: " + str(max_row_count) + ", Total Column: " + str(max_col_count))

# Read the columns value
# for i in range(1, max_col_count+1):
#     cell_obj = s.cell(1, i)
#     print(cell_obj.value)
# Read the row values
for i in range(1, max_row_count+1):
    for j in range(1, max_col_count+1):
        cell_obj = s.cell(i, j)
        print(cell_obj.value, end=" ")
    print()
