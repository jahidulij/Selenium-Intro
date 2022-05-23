import openpyxl

# Load workbook with path
path = "C:\\Users\\Asus\\workspace_python\\seleniumIntro\\Screenshots\\TestData.xlsx"
wb = openpyxl.load_workbook(path)
# Identify active worksheet/ worksheet choose
s = wb.active
# Get max row & column
max_row_count = s.max_row
max_col_count = s.max_column
print("Total row: " + str(max_row_count) + ", Total Column: " + str(max_col_count))
# Sheet title
sheet_title = s.title
print(sheet_title)

s.cell(row=1, column=4).value = "output"

# Write the values in excel
for i in range(2, max_row_count+1):
    if s.cell(i, 3).value == "Valid":
        s.cell(row=i, column=4).value = "Login successful"
    else:
        s.cell(row=i, column=4).value = "Login unsuccessful"

wb.save(path)
